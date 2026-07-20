"""Parse 2020 CPH Table B (population by province/city/municipality for the
2000, 2010, 2015, 2020 censuses) into a tidy municipal panel, resolved to PSGC.

Structure of each region sheet:
    - header rows; the row of census reference dates identifies data columns
    - province rows carry =SUM(...) formulas in the population columns
    - municipality rows carry literal numbers
    - region row is the first formula row

Output: data/interim/census_pop_muni.csv
    psgc10, name_src, province_src, region_sheet, pop2000, pop2010, pop2015,
    pop2020, match_method, flags
Unresolved rows are kept with psgc10="" and listed in the build log — never
silently dropped.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from psgc import resolve  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
INTERIM = ROOT / "data" / "interim"
TABLE_B = RAW / "census" / "2020cph_tableB_pop_growth_muni.xlsx"

POP_COLS = {2000: 3, 2010: 4, 2015: 5, 2020: 6}  # 1-indexed worksheet columns C..F


def parse_sheet(ws, sheet_name: str) -> list[dict]:
    rows = []
    province = None
    seen_region_header = False
    for r in ws.iter_rows(min_row=1, values_only=True):
        name = r[0]
        if not isinstance(name, str) or not name.strip():
            continue
        if name.strip().startswith(("B.", "REGION, PROVINCE", "Note", "Source", "1/", "2/", "3/")):
            continue
        c = r[2]  # first population column (2000)
        c2020 = r[5]
        is_formula = any(isinstance(v, str) and v.startswith("=") for v in (r[2], r[3], r[4], r[5]))
        if is_formula:
            if not seen_region_header:
                seen_region_header = True  # region total row
            else:
                province = name.strip()
            continue
        vals = {}
        numeric_any = False
        for yr, ci in POP_COLS.items():
            v = r[ci - 1]
            if isinstance(v, (int, float)):
                vals[f"pop{yr}"] = int(v)
                numeric_any = True
            else:
                vals[f"pop{yr}"] = None
        if not numeric_any:
            continue
        rows.append({
            "name_src": name.strip(),
            "province_src": province,
            "region_sheet": sheet_name,
            **vals,
        })
    return rows


def main() -> None:
    wb = openpyxl.load_workbook(TABLE_B, read_only=True)
    allrows = []
    for s in wb.sheetnames:
        allrows += parse_sheet(wb[s], s)
    df = pd.DataFrame(allrows)
    munis = pd.read_csv(INTERIM / "psgc_municipalities.csv", dtype={"psgc10": str})
    res = resolve(df, munis)
    INTERIM.mkdir(parents=True, exist_ok=True)
    res.to_csv(INTERIM / "census_pop_muni.csv", index=False)
    n_un = (res.match_method == "UNRESOLVED").sum()
    print(f"rows: {len(res)}  unresolved: {n_un}")
    if n_un:
        print(res[res.match_method == "UNRESOLVED"][["name_src", "province_src", "region_sheet"]].to_string())
    # sanity: national totals
    for yr in (2000, 2010, 2015, 2020):
        print(yr, int(res[f"pop{yr}"].sum(skipna=True)))


if __name__ == "__main__":
    main()
