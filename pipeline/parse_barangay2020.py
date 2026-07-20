"""Barangay-level 2020 census populations (total + household), all 17 regions,
from the archived per-region 2020 CPH files ("Total Population, Household
Population, and Number of Households by Province, City, and Municipality and
Barangay: as of 01 May 2020").

Sheet grammar (per province sheet): PROVINCE row (caps, = sheet name), then
municipality rows (caps), then that municipality's barangays (mixed case).
NCR uses one flat sheet with city rows in caps. BARMM includes an "Interim
Province" sheet (the future SGA barangays).

Barangay PSGC resolution: municipality context is resolved with the standard
muni resolver; barangays are then matched within the municipality's PSGC
block (codes sharing the 7-digit prefix). Unmatched barangays are kept with
an empty code and flagged — never dropped.

Output: data/clean/population_barangay_2020.csv
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from psgc import load_psgc, normalize_name, strip_parens, resolve  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SRCDIR = ROOT / "data" / "raw" / "census" / "barangay2020"
CLEAN, INTERIM = ROOT / "data" / "clean", ROOT / "data" / "interim"

FILES = [  # (file, region label); NCR_0 has the barangay sheet for NCR
    ("NCR_0.xlsx", "National Capital Region (NCR)"),
    ("CAR_1.xlsx", "Cordillera Administrative Region (CAR)"),
    ("Region_I.xlsx", "Region I (Ilocos Region)"),
    ("Region_II.xlsx", "Region II (Cagayan Valley)"),
    ("Region_III.xlsx", "Region III (Central Luzon)"),
    ("Region_IV-A.xlsx", "Region IV-A (CALABARZON)"),
    ("MIMAROPA.xlsx", "MIMAROPA Region"),
    ("Region_V.xlsx", "Region V (Bicol Region)"),
    ("Region_VI.xlsx", "Region VI (Western Visayas)"),
    ("Region_VII.xlsx", "Region VII (Central Visayas)"),
    ("Region_VIII_1.xlsx", "Region VIII (Eastern Visayas)"),
    ("Region_IX.xlsx", "Region IX (Zamboanga Peninsula)"),
    ("Region_X_1.xlsx", "Region X (Northern Mindanao)"),
    ("Region_XI.xlsx", "Region XI (Davao Region)"),
    ("Region_XII.xlsx", "Region XII (SOCCSKSARGEN)"),
    ("Region_XIII_(CARAGA).xlsx", "Region XIII (Caraga)"),
    ("BARMM.xlsx", "Bangsamoro Autonomous Region In Muslim Mindanao (BARMM)"),
]


def is_caps(s: str) -> bool:
    """Caps test on the paren-stripped name: context rows print annotations
    like "CITY OF CEBU (Capital)" whose parenthetical is mixed-case."""
    core = strip_parens(s)
    letters = [c for c in core if c.isalpha()]
    return bool(letters) and all(c.upper() == c for c in letters)


def norm_bgy(s: str) -> str:
    """Keep parenthetical CONTENT (it disambiguates: 'San Roque (Pob.)' vs
    'San Roque'); only the parens themselves are dropped."""
    s = str(s).replace("(", " ").replace(")", " ")
    s = normalize_name(s)
    s = re.sub(r"\bpob\b", "poblacion", s)
    s = re.sub(r"\bbgy\b|\bbrgy\b|\bbarangay\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def walk_file(path: Path, region: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = []
    sheets = wb.sheetnames
    for sheet in sheets:
        low = sheet.lower()
        if low.startswith(("region", "car", "barmm", "mimaropa", "ncr by")) and len(sheets) > 1:
            continue  # region-summary sheet (muni totals; we get those elsewhere)
        if sheet == "by City & Mun":
            continue
        ws = wb[sheet]
        province = None
        muni = None
        for r in ws.iter_rows(min_row=1, values_only=True):
            name = r[0]
            if not isinstance(name, str) or not name.strip():
                continue
            name = name.strip()
            if name.upper().startswith(("TABLE", "(TOTAL", "PROVINCE", "AND BARANGAY",
                                        "CITY, AND", "SOURCE", "NOTE")):
                continue
            tot, hh = r[1], r[2]
            has_val = isinstance(tot, (int, float))
            if is_caps(name):
                if re.search(r"(REGION|CORDILLERA|BANGSAMORO|CARAGA\)|MIMAROPA|CAPITAL)",
                             strip_parens(name), re.I) and "CITY" not in name.upper():
                    continue
                if province is None:
                    province = name
                    continue
                muni = name
                continue
            if not has_val:
                continue
            rows.append({"region": region, "province_sheet": sheet,
                         "province_src": province or sheet, "muni_src": muni,
                         "barangay_src": name,
                         "pop2020_total": int(tot),
                         "pop2020_household": int(hh) if isinstance(hh, (int, float)) else None})
    return rows


def main() -> None:
    allrows = []
    for f, region in FILES:
        p = SRCDIR / f
        rows = walk_file(p, region)
        allrows += rows
        print(f"{f}: {len(rows)} barangay rows")
    df = pd.DataFrame(allrows)

    # resolve municipality context (NCR: city rows are the caps context)
    munis = pd.read_csv(INTERIM / "psgc_municipalities.csv", dtype=str)
    ctx = df[["muni_src", "province_src", "region"]].drop_duplicates().copy()
    ctx["name_src"] = ctx.muni_src
    ctx.loc[ctx.region.str.contains("NCR"), "province_src"] = "National Capital Region (NCR)"
    resolved = resolve(ctx.dropna(subset=["name_src"]), munis)
    muni_code = {(r.muni_src, r.province_src, r.region): r.psgc10
                 for r in resolved.itertuples() if r.match_method != "UNRESOLVED"}
    un_m = resolved[resolved.match_method == "UNRESOLVED"]
    if len(un_m):
        print("unresolved municipality contexts:", len(un_m))
        print(un_m[["name_src", "province_src"]].head(10).to_string())

    # Manila's sub-municipal districts: caps contexts on the NCR sheet that are
    # not municipalities — match against PSGC SubMun entries (their codes are
    # the prefix under which Manila's barangays sit)
    psgc = load_psgc()
    submun = psgc[psgc.level == "SubMun"].copy()
    submun_by_norm = {}
    for t in submun.itertuples():
        submun_by_norm.setdefault(normalize_name(strip_parens(t.name)), t.psgc10)
    # source prints "TONDO I/II" for Tondo I / Tondo II — both share barangays
    # only via their own codes; treat the combined label as Tondo I's block
    # plus fallback matching across all Manila blocks below.
    MANILA = "1380600000"

    bgy = psgc[psgc.level == "Bgy"].copy()
    bgy["mun7"] = bgy.psgc10.str[:7]
    namecol = "name" if "name" in bgy.columns else "Name"
    bgy["key"] = bgy[namecol].map(norm_bgy)
    bgy["key_stripped"] = bgy[namecol].map(lambda s: norm_bgy(strip_parens(str(s))))
    index: dict[tuple[str, str], list[str]] = {}
    index_stripped: dict[tuple[str, str], list[str]] = {}
    manila_index: dict[str, list[str]] = {}
    for t in bgy.itertuples():
        index.setdefault((t.mun7, t.key), []).append(t.psgc10)
        index_stripped.setdefault((t.mun7, t.key_stripped), []).append(t.psgc10)
        if t.psgc10.startswith(MANILA[:5]):
            manila_index.setdefault(t.key, []).append(t.psgc10)

    out = []
    n_ok = n_amb = n_miss = 0
    for t in df.itertuples():
        mcode = muni_code.get((t.muni_src, "National Capital Region (NCR)" if "NCR" in t.region
                               else t.province_src, t.region))
        rec = {"region": t.region, "province": t.province_src, "municipality": t.muni_src,
               "muni_psgc10": mcode or "", "barangay_src": t.barangay_src,
               "pop2020_total": t.pop2020_total, "pop2020_household": t.pop2020_household,
               "psgc10": "", "flag": ""}
        if mcode:
            k = norm_bgy(t.barangay_src)
            cands = index.get((mcode[:7], k), [])
            method = "exact"
            if len(cands) != 1:
                c2 = index_stripped.get((mcode[:7], norm_bgy(strip_parens(t.barangay_src))), [])
                if len(c2) == 1:
                    cands, method = c2, "paren-stripped"
            if len(cands) == 1:
                rec["psgc10"] = cands[0]; n_ok += 1
                if method != "exact":
                    rec["flag"] = "matched after dropping parenthetical"
            elif len(cands) > 1:
                rec["flag"] = "ambiguous barangay name within municipality"; n_amb += 1
            else:
                rec["flag"] = "barangay name not matched to PSGC"; n_miss += 1
        elif isinstance(t.muni_src, str) and (
                normalize_name(strip_parens(t.muni_src)) in submun_by_norm
                or "NCR" in t.region):
            # Manila SubMun context: match across Manila's barangay blocks
            cands = manila_index.get(norm_bgy(t.barangay_src), [])
            if len(cands) == 1:
                rec["psgc10"] = cands[0]; rec["muni_psgc10"] = MANILA
                rec["flag"] = "Manila (SubMun context)"; n_ok += 1
            else:
                rec["flag"] = "Manila SubMun: barangay not uniquely matched"; n_miss += 1
        else:
            rec["flag"] = "municipality context unresolved"
        out.append(rec)
    res = pd.DataFrame(out)
    res.to_csv(CLEAN / "population_barangay_2020.csv", index=False)
    print(f"barangays: {len(res)} | psgc-matched: {n_ok} ({n_ok/len(res)*100:.1f}%) "
          f"| ambiguous: {n_amb} | unmatched: {n_miss}")
    print("national total:", int(res.pop2020_total.sum()))


if __name__ == "__main__":
    main()
